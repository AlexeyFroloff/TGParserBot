import openpyxl  #для Xlsx
import requests
from openpyxl.styles import (Font, Alignment, Border, Side)
import datetime
from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.dispatcher.filters import Text
from aiogram.utils import executor
from datetime import timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options as FirefoxOptions
import asyncio
import aioschedule

if True:
    slmonth = {1:"January", 2:"February", 3:"March", 4:"April", 5:"May", 6:"June", 7:"July", 8:"August", 9:"September", 10: "October", 11:"November", 12: "December", 0: "December"}
    slmonthru = {1:"Январь", 2:"Февраль", 3:"Март", 4:"Апрель", 5:"Май", 6:"Июнь", 7:"Июль", 8:"Август", 9:"Сентябрь", 10: "Октябрь", 11:"Ноябрь", 12: "Декабрь", 0: "Декабрь"}

    sp = str("A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z").split(', ')
    letters = sp
    for i in range(26):  # letters - список горизонтальных индексов
        for j in range(26):
            letters.append(str(sp[i]) + str(sp[j]))

    now = datetime.datetime.now()
    my_cookies = ""
    s = requests.Session()

def get_cookies():
    global my_cookies

    options = webdriver.FirefoxOptions()
    options.headless = True

    browser = webdriver.Firefox(options = options)
    browser.get("https://online.unicum.ru/n/user/")

    #username = 
    #password = 

    elem1 = browser.find_element(By.XPATH,
                                 "/html/body/table[2]/tbody/tr/td/form/table/tbody/tr[1]/td[3]/table/tbody/tr[1]/td[2]/input")
    elem1.send_keys(username)

    elem2 = browser.find_element(By.XPATH,
                                 "/html/body/table[2]/tbody/tr/td/form/table/tbody/tr[1]/td[3]/table/tbody/tr[3]/td[2]/input")
    elem2.send_keys(password)

    elem3 = browser.find_element(By.XPATH, "/html/body/table[2]/tbody/tr/td/form/table/tbody/tr[3]/td[3]/input")
    elem3.click()

    my_cookies = browser.get_cookies()[1]["value"]  # nvmc_login

    print("cookies have been updated " + my_cookies + " " + str(datetime.datetime.now()))

    browser.close()

    return my_cookies

#get_data returns only "objects" - object's information data {}
def get_data(start_d, start_m, start_y, end_d, end_m, end_y):
    global my_cookies
    print(datetime.datetime.now())

    #url = {cсылки на объекты}

    s = requests.Session()


    def __main__(start_d, start_m, start_y, end_d, end_m, end_y):#вводим желаемый период

        headers = {
            "day": start_d,
            "month": start_m,
            "year": start_y,
            "day1": end_d,
            "month1": end_m,
            "year1": end_y,
        }
        return headers

    datas = __main__(start_d, start_m, start_y, end_d, end_m, end_y)

    sumnumber = 0
    summoney = 0

    objects = {}
    for i in range(1, 55):
        try:
            loging = s.post(url[i], cookies={"nvmc_login": my_cookies, "nvmc_root": "/n/"}, data=datas)
            if loging.status_code != 200:
                s = requests.Session()
                get_cookies()
                loging = s.post(url[i], cookies={"nvmc_login": my_cookies, "nvmc_root": "/n/"}, data=datas)

            if loging.status_code == 200:
                sl = str(loging.text).split('TR')
                el = sl[len(sl) - 4]
                el1 = el.split("SIZE=")[-2]
                el2 = el1[el1.index(">") + 1:el1.index("<")].split("&nbsp;")
                number, money = int(el2[0]), int(round(float(el2[2]), 0))
                sumnumber += number
                summoney += money

                # ищем location в коде
                location = ' '.join(
                    loging.text[loging.text.index('Адрес:') + 7:loging.text.index('<BR')].split("&nbsp;"))

                if type(number) == int and type(money) == int:
                    objects[i] = [number, money, location]

                print(location, ": ", number, "; ", money, sep="")
                i += 1
        except:
            print("Cookie or Network Error!")


                                                #
    print("Итого:", sumnumber, summoney)

    return objects

def write_month(data, sheet, auto, i, start=1, u=0, k=1):
    i+=1
    ws = auto.active
    s = sheet
    global letters
    print(data)

    for j in range(55):
        if s["A" + str(j + 2)].value in data.keys():                    # заполнение данными за месяц

            s["B" + str(j + 2)] = data[s["A" + str(j + 2)].value][2]
            s["B" + str(j + 2)] = data[s["A" + str(j + 2)].value][2]


            s[letters[i * 3 - 2 - u] + str(j + 2)] = data[s["A" + str(j + 2)].value][0]
            C = ws[str(letters[i * 3 - 2 - u]) + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            s[letters[i * 3 - 1 - u] + str(j + 2)] = data[s["A" + str(j + 2)].value][1]
            C = ws[letters[i * 3 - 1 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            if k != 0:
                s[letters[i * 3 - u] + str(j + 2)] = (
                            "=" + letters[i * 3 - 1 - u] + str(j + 2) + " *0.75" + "-D" + str(j + 2))
            elif k == 0:
                s[letters[i * 3 - u] + str(j + 2)] = ("=" + letters[i * 3 - 1 - u] + str(j + 2) + " *0.75")
            C = ws[letters[i * 3 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            s[letters[i * 3 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

        elif s["B" + str(j + 2)].value == "Итого, Адлер:":  # Итогошки
            itogadler = j + 2
            value = "SUM(" + letters[i * 3 - 2 - u] + "2:" + letters[i * 3 - 2 - u] + str(j + 1) + ")"
            s[letters[i * 3 - 2 - u] + str(j + 2)] = "=IF(" + str(value) + '=0,"",' + str(value) + ")"
            # print(s[letters[i*3-2-u] + str(j+2)].value)
            s[letters[i * 3 - 2 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - 2 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            s[letters[i * 3 - 1 - u] + str(j + 2)] = (
                        "=SUM(" + letters[i * 3 - 1 - u] + "2:" + letters[i * 3 - 1 - u] + str(j + 1) + ")")
            s[letters[i * 3 - 1 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - 1 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            s[letters[i * 3 - u] + str(j + 2)] = (
                        "=SUM(" + letters[i * 3 - u] + "2:" + letters[i * 3 - u] + str(j + 1) + ")")
            s[letters[i * 3 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

        elif s["B" + str(j + 2)].value == "Итого, Москва:":
            itogmoscow = j + 2
            s[letters[i * 3 - 2 - u] + str(j + 2)] = (
                        "=SUM(" + letters[i * 3 - 2 - u] + str(itogadler + 1) + ":" + letters[i * 3 - 2 - u] + str(
                    j + 1) + ")")
            s[letters[i * 3 - 2 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - 2 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            s[letters[i * 3 - 1 - u] + str(j + 2)] = (
                        "=SUM(" + letters[i * 3 - 1 - u] + str(itogadler + 1) + ":" + letters[i * 3 - 1 - u] + str(
                    j + 1) + ")")
            s[letters[i * 3 - 1 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - 1 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            s[letters[i * 3 - u] + str(j + 2)] = (
                        "=SUM(" + letters[i * 3 - u] + str(itogadler + 1) + ":" + letters[i * 3 - u] + str(j + 1) + ")")
            s[letters[i * 3 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

        elif s["B" + str(j + 2)].value == "Итого:":
            s[letters[i * 3 - 2 - u] + str(j + 2)] = (
                        "=" + letters[i * 3 - 2 - u] + str(itogadler) + "+" + letters[i * 3 - 2 - u] + str(itogmoscow))
            s[letters[i * 3 - 2 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - 2 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            s[letters[i * 3 - 1 - u] + str(j + 2)] = (
                        "=" + letters[i * 3 - 1 - u] + str(itogadler) + "+" + letters[i * 3 - 1 - u] + str(itogmoscow))
            s[letters[i * 3 - 1 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - 1 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

            s[letters[i * 3 - u] + str(j + 2)] = (
                        "=" + letters[i * 3 - u] + str(itogadler) + "+" + letters[i * 3 - u] + str(itogmoscow))
            s[letters[i * 3 - u] + str(j + 2)].font = Font(bold=True)
            C = ws[letters[i * 3 - u] + str(j + 2)]
            C.alignment = Alignment(horizontal='center')

        else:
            s[letters[i * 3 - u] + str(j + 2)] = ""
            s[letters[i * 3 - u] + str(j + 2)] = ""
            s[letters[i * 3 - u] + str(j + 2)] = ""

#successfully shows last month statistics, using the sample
def write_last_month(file = "just_a_sample.xlsx", sheet_name = "Лист1"):
    now = datetime.datetime.now()
    month = now.month - 1
    if month == 1:
        year = now.year - 1
    else:
        year = now.year

    auto = openpyxl.load_workbook(file)
    sheet = auto[sheet_name]
    ws = auto.active
    ws.merge_cells('E1:G1')
    ws["E1"] = slmonthru[month] + " " + str(now.year)
    C = ws["E1"]
    C.alignment = Alignment(horizontal='center')
    save_filename = "Stat" + slmonth[month]+ ".xlsx"
    if str(month) in {"4", "6", "9", "11"}:
        end_d = "30"
    elif str(month) in {2}:
        end_d = "28"
    else:
        end_d = "31"
    data = get_data(1, month, year, end_d, month, year)
    write_month(data, sheet, auto, i = 1)
    auto.save(save_filename)
    return save_filename

def write_current_month(file = "just_a_sample.xlsx", sheet_name = "Лист1"):
    now = datetime.datetime.now()
    month = now.month
    year = now.year

    auto = openpyxl.load_workbook(file)
    sheet = auto[sheet_name]
    ws = auto.active
    ws.merge_cells('E1:G1')
    ws["E1"] = slmonthru[month] + " " + str(now.year)
    C = ws["E1"]
    C.alignment = Alignment(horizontal='center')
    save_filename = "CurrentStat" + slmonth[month]+ ".xlsx"
    if str(month) in {"4", "6", "9", "11"}:
        end_d = "30"
    elif str(month) in {2}:
        end_d = "28"
    else:
        end_d = "31"
    data = get_data(1, month, year, end_d, month, year)
    write_month(data, sheet, auto, i = 1)
    auto.save(save_filename)
    return save_filename

def write_day(a, file = "just_a_sample_today.xlsx", sheet_name = "Лист1"):
    now = datetime.datetime.now()
    yesterday = now - timedelta(days=1)

    auto = openpyxl.load_workbook(file)
    sheet = auto[sheet_name]
    ws = auto.active
    ws.merge_cells('C1:E1')
    ws["C1"] = slmonthru[now.month] + " " + str(now.year)
    C = ws["C1"]
    C.alignment = Alignment(horizontal='center')
    if a == "to":
        save_filename = "TodayStat_" + str(now.day) + "_" + str(now.month) + "_" + str(now.year) + ".xlsx"
        data = get_data(now.day, now.month, now.year, now.day, now.month, now.year)
        write_month(data, sheet, auto, 1, now.day, 2, 0)
        auto.save(save_filename)
        return save_filename
    elif a == "yester":
        save_filename = "YesterdayStat_" + str(yesterday.day) + "_" + str(yesterday.month) + "_" + str(yesterday.year) + ".xlsx"
        data = get_data(yesterday.day, yesterday.month, yesterday.year, yesterday.day, yesterday.month, yesterday.year)
        write_month(data, sheet, auto, 1, yesterday.day, 2, 0)
        auto.save(save_filename)
        return save_filename

def statistics(file = "just_a_sample.xlsx", sheet_name = "Лист1"):
    now = datetime.datetime.now()
    time = datetime.date(2022, 1, 1).month
    year = datetime.date(2022, 1, 1).year
    number = now.month - time + 1 + (now.year - year)*12            #количество месяцев
    auto = openpyxl.load_workbook(file)
    sheet = auto[sheet_name]
    ws = auto.active
    for t in range(number):
        ws.merge_cells(str(letters[3*t+4]) + "1:" + str(letters[3*t+6]) + "1")
    for j in range(number):
        month = (time + j-1)%12+1
        year = 2022 + (j)//12
        ws[(letters[3*j+4]+"1")] = slmonthru[month] + " " + str(year)
        C = ws[(letters[3*j+4]+"1")]
        C.alignment = Alignment(horizontal='center')
        if str(month) in {"4", "6", "9", "11"}:
            end_d = "30"
        elif str(month) in {2}:
            end_d = "28"
        else:
            end_d = "31"
        data = get_data(1, month, year, end_d, month, year)
        write_month(data, sheet, auto, i=j+1)
        save_filename = "АНАЛИЗ_ПО_МЕСЯЦАМ_АППАРАТЫ" + "_" + str(now.day) + "_" + str(now.month) + "_" + str(now.year) + ".xlsx"
        auto.save(save_filename)

    return save_filename

if True:
    bot = Bot(token = '')
    dp = Dispatcher(bot)
    chat_id = 0#bot information #bot information
    print("woked up")

def errors_data():
    global info
    global my_cookies
    global s
    c = 0
    sp = []
    spsl = []
    try:
        print(1)
        l = s.post("https://online.unicum.ru/n/vmonitor.html", cookies={"nvmc_login": my_cookies, "nvmc_root": "/n/"})
        print(2)
        if l.status_code != 200:
            s = requests.Session()
            get_cookies()
            l = s.post("https://online.unicum.ru/n/vmonitor.html", cookies={"nvmc_login": my_cookies, "nvmc_root": "/n/"})
        html = str(l.text)
        html = html[html.index("TABLE BORDER=1") + 19:]

        sp = html.split("<TR>")

    except:
        print("Cookie or Network Error")



    for el in sp:
        if c != 0:
            sl = {}
            string = el.split("</TD><TD>")
            for i in range(len(string)):
                el1 = string[i]
                if i == 4:
                    el1 = el1[20:]
                    el1 = el1.split("</FONT> ")
                    sl["Статус"] = el1[0]
                    sl["Время"] = el1[1]
                elif "ALT" in el1 and i != 15:
                    status_alt = el1[el1.index("ALT")+4:el1.index("</A>")-10]
                    if i == 5:
                        sl["Контроллер"] = status_alt
                    elif i == 6:
                        sl["ТА"] = status_alt
                    elif i == 8:
                        sl["Монетоприемник"] = status_alt
                    elif i == 9:
                        sl["Банкнотоприемник"] = status_alt
                    elif i == 10:
                        sl["КРТ"] = status_alt
                    elif i == 11:
                        sl["ККТ"] = status_alt
                    elif i == 12:
                        sl["Н"] = status_alt
                    elif i == 13:
                        sl["П"] = status_alt
                elif "Контроллер" not in list(sl.keys()) and i == 5:
                    sl["Контроллер"] = "-"
                elif "ТА" not in list(sl.keys()) and i == 6:
                    sl["ТА"] = "-"
                elif "Монетоприемник" not in list(sl.keys()) and i == 8:
                    sl["Монетоприемник"] = "-"
                elif "Банкнотоприемник" not in list(sl.keys()) and i == 9:
                    sl["Банкнотоприемник"] = "-"
                elif "КРТ" not in list(sl.keys()) and i == 10:
                    sl["КРТ"] = "-"
                elif "ККТ" not in list(sl.keys()) and i == 11:
                    sl["ККТ"] = "-"
                elif "Н" not in list(sl.keys()) and i == 12:
                    sl["Н"] = "-"
                elif "П" not in list(sl.keys()) and i == 13:
                    sl["П"] = "-"
                elif i == 1:
                    el2 = string[i]
                    url = el2[el2.index("HREF=")+5:el2.index(">")]
                    url = url[url.index("?")+1:]
                    sl["url"] = url
                    serial_number = el2[el2.index(">")+1:el2.index("/A")]
                    sl["Серийный номер"] = serial_number[:-1]
                elif i == 2:
                    el2 = string[i]
                    address = ' '.join(el2.split("&nbsp;"))
                    sl["Адрес"] = address
                elif i == 3:
                    el2 = string[i]
                    sl["Маршрут"] = el2
                elif i == 4:
                    el2 = string[i]
                elif i == 7:
                    el2 = string[i]
                    o = []
                    if "FONT" in el2:
                        el2 = el2.split("<")
                    for j in range(len(el2)):
                        if "FONT COLOR" in el2[j]:
                            number = el2[j][el2[j].index(">")+1:]
                            o.append(number)
                        elif "/" in el2[j] and "FONT" not in el2[j]:
                            number = el2[j][:el2[j].index("/")]
                            o.append(number)
                    sl["Сдача"] = '/'.join(o)
                elif i == 8:
                    el2 = string[i]
                    el2 = el2[el2.index(">")+1:]
                    el2 = el2[:el2.index("<")]
                    sl["Монетоприемник"] = el2

                elif i == 14:
                    el2 = string[i]
                    sl["Продаж"] = el2
                elif i == 15:
                    el2 = string[i]
                    if "ALT" not in el2:
                        sl["∑пр1"] = '"OK"'
                        sl["∑пр2"] = el2
                    else:
                        st = el2[el2.index("ALT")+4:el2.index("</A>")-10]
                        numb = el2[el2.rindex(">")+1:]
                        sl["∑пр1"] = st
                        sl["∑пр2"] = numb
                elif i == 16:
                    el2 = string[i]
                    sl["Обслуживание"] = el2
                elif i == 17:
                    el2 = string[i]
                    el2 = el2[:el2.index("<")]
                    sl["След. обслуж."] = el2
            sl["Подробнее"] = (str("https://online.unicum.ru/n/curerrors.html?V") + str(sl["url"]))
            sl["Номер"] = [c]
            spsl.append(sl)
        c += 1
    return spsl

info = errors_data()

def get_message():
    global info

    text = []
    spsl = errors_data()
    if spsl != []:
        keys = list(spsl[0].keys())
    else:
        keys = []
    changes = []
    allchanges = []
    for i in range(len(spsl)):
        sl = spsl[i]
        for el in keys:
           #print("!!!!!!!!!!", el, i, len(info), info, "!!!!!!!!!!!!!!!")
            if el != "Время" and sl[el] != info[i][el]:
                allchanges.append([sl["Адрес"], sl["Время"], "было:", info[i][el], "стало:", sl[el], sl["Номер"]])
            if el in ["Статус", "Контроллер", "ТА", "Сдача", 'Монетоприемник', 'Банкнотоприемник', 'КРТ', 'ККТ', 'Н', 'П',
                      '∑пр1', 'Обслуживание'] and sl[el] != info[i][el]:
                changes.append([sl["Адрес"], sl["Время"], "было:", info[i][el], "стало:", sl[el], sl["Номер"], el])


    for el in changes:
        #print(el)
       # print("!!!!!!!!", el[-1], str(el[3]), str(el[5]))
        if el[-1] == "Статус":
            text.append("Статус объекта " + str(el[0]) + " изменился в " + str(el[1]) + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0])-1]["Подробнее"]))


        elif el[-1] == "Контроллер" and str(el[3]) == '"OK"' and str(el[5]) == '"ОШИБКА"':
            text.append(str(el[0]) + ": в контроллере в " + str(el[1]) + " произошла ошибка." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0])-1]["Подробнее"]))
        elif el[-1] == "Контроллер" and str(el[3]) == '"ОШИБКА"' and str(el[5]) == '"OK"':
            text.append(str(el[0]) + ": контроллер заработал в " + str(el[1]) + "." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0])-1]["Подробнее"]))


        elif el[-1] == "ТА" and str(el[3]) == '"OK"' and str(el[5]) == '"ОШИБКА"':
            text.append(str(el[0]) + ": в ТА в " + str(el[1]) + " произошла ошибка." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))
        elif el[-1] == "ТА" and str(el[3]) == '"ОШИБКА"' and str(el[5]) == '"OK"':
            text.append(str(el[0]) + ": ТА заработал в " + str(el[1]) + "." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))


        elif el[-1] == "Монетоприемник" and str(el[3]) == '"OK"' and str(el[5]) == '"ОШИБКА"':
            text.append(str(el[0]) + ": в монетоприемнике в " + str(el[1]) + " произошла ошибка." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))
        elif el[-1] == "Монетоприемник" and str(el[3]) == '"ОШИБКА"' and str(el[5]) == '"OK"':
            text.append(str(el[0]) + ": монетоприемник заработал в " + str(el[1]) + "." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))


        elif el[-1] == "Банкнотоприемник" and str(el[3]) == '"OK"' and str(el[5]) == '"ОШИБКА"':
            text.append(str(el[0]) + ": в банкнотоприемнике в " + str(el[1]) + " произошла ошибка." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))
        elif el[-1] == "Банкнотоприемник" and str(el[3]) == '"ОШИБКА"' and str(el[5]) == '"OK"':
            text.append(str(el[0]) + ": банкнотоприемник заработал в " + str(el[1]) + "." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))


        elif el[-1] == "КРТ" and str(el[3]) == '"OK"' and str(el[5]) == '"ОШИБКА"':
            text.append(str(el[0]) + ": в считывателе карт в " + str(el[1]) + " произошла ошибка." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))
        elif el[-1] == "КРТ" and str(el[3]) == '"ОШИБКА"' and str(el[5]) == '"OK"':
            text.append(str(el[0]) + ": Работа считывателя карт восстановлена - " + str(el[1]) + "." + '\n' +
                  "Текущий статус: " + str(el[5]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))


        elif el[-1] == "Обслуживание":
            text.append(str(el[0]) + ": Аппарат был обслужен " + str(el[1]) + "." + '\n' +
                  "Предыдущее обслуживание: " + str(el[3]) + '\n' +
                  "Подробнее: " + str(spsl[int(el[6][0]) - 1]["Подробнее"]))

    info = spsl.copy()

    return text

def get_errors_data():
    errors = errors_data()

    error_message = []

    ok_statuses = ['"ОК"', "ОК", "-", " ", "", "OK", '"OK"']
    bot_mes = ""
    for machine in errors:
        # print(machine)
        mes = ""
        if machine["Статус"] != "OnLine":
            mes += "Статус: " + machine["Статус"] + "\n"
        for el in ['Контроллер', 'ТА', 'Сдача', 'Монетоприемник', 'Банкнотоприемник', 'КРТ', 'ККТ', 'П']:
            if machine[el] not in ok_statuses:
                mes += el + ": " + machine[el] + "\n"
        if mes != "":
            bot_mes += machine["Адрес"] + ":\n" + mes + "\n" * 2
    file = open("Errors.txt", "w")
    file.write(bot_mes)
    file.close()
    return file

@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    #global chat_id
    #global chats
    chat_id = message.chat.id
    chats = []
    for el in open("chats.txt", "r").read().split():
        chats.append(el)
    if str(chat_id) not in chats:
        file = open("chats.txt", "w")
        for el in chats:
            file.write(str(el) + " ")
        file.write(str(chat_id))
        file.close()#("chats.txt")
    #chats.append(chat_id)
    try:
        kb = [
            [
                types.KeyboardButton(text="Статистика за текущий месяц"),
                types.KeyboardButton(text="Статистика за прошлый месяц"),
            ],
            [
                types.KeyboardButton(text="Статистика за сегодня"),
                types.KeyboardButton(text="Статистика за вчера"),
            ],
            [
                types.KeyboardButton(text="Общая статистика")
            ],
            [
                types.KeyboardButton(text="Текущие ошибки")
            ],
        ]
        keyboard = types.ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)
        await message.reply("Привет!\nЯ Лера.\nЧто бы вы хотели получить?", reply_markup=keyboard)
    except:
        True

current_month = now.month
@dp.message_handler(Text("Статистика за текущий месяц"))
async def with_puree(message: types.Document):
    try:
        await message.reply("Минутку...")
        file = write_current_month()
        await message.reply_document(open(file, "rb"))
    except:
        True

last_month = slmonthru[(current_month+11)%12]
@dp.message_handler(lambda message: message.text == "Статистика за прошлый месяц")
async def without_puree(message: types.Message):
    try:
        await message.reply("Минутку...")
        file = write_last_month()
        await message.reply_document(open(file, "rb"))
    except:
        True

@dp.message_handler(Text("Общая статистика"))
async def with_puree(message: types.Document):
    try:
        await message.reply("Минутку...")
        file = statistics()
        await message.reply_document(open(file, "rb"))
    except:
        print("There")

@dp.message_handler(Text("Статистика за сегодня"))
async def with_puree(message: types.Document):
    try:
        await message.reply("Минутку...")
        file = write_day("to")
        await message.reply_document(open(file, "rb"))
    except:
        True

@dp.message_handler(Text("Статистика за вчера"))
async def with_puree(message: types.Document):
    try:
        await message.reply("Минутку...")
        file = write_day("yester")
        await message.reply_document(open(file, "rb"))
    except:
        True

async def sender():
    messages = get_message()
    print(str(datetime.datetime.now()), messages)
    if messages != []:
        for message in messages:
            print(message)
            for chat in open("chats.txt", "r").read().split():
                await bot.send_message(chat_id=int(chat), text=message)
            print("OK")
    else:
        print(str(datetime.datetime.now()) + "- ошибок не обнаружено")

@dp.message_handler(Text("Текущие ошибки"))
async def with_puree(message: types.Document):
    await message.reply("Минутку...")
    await sender()
    try:
        file = get_errors_data()
        await message.reply_document(open("Errors.txt", "rb"))
        await message.reply("Подбробнее: https://online.unicum.ru/n/vmonitor.html")
    except:
        True

async def scheduler():
    aioschedule.every(400).seconds.do(sender)
    while True:
        await aioschedule.run_pending()
        await asyncio.sleep(200)

async def on_startup(dp):
    asyncio.create_task(scheduler())

if __name__ == '__main__':
    executor.start_polling(dp, on_startup=on_startup)
    executor.start_polling(dp, skip_updates=True)
